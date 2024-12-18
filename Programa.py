import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os

# Ruta dinámica del archivo Excel
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "Inventario.xlsx")

# Verificar existencia del archivo Excel
def verificar_archivo_excel():
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("Error", f"El archivo {EXCEL_FILE} no se encuentra en el directorio actual.")
        exit()

# Cargar datos de Excel
def cargar_catalogo():
    df = pd.read_excel(EXCEL_FILE, sheet_name="Productos VALMEX")
    try:
        df['PF (Precio Final)'] = pd.to_numeric(df['PF (Precio Final)'], errors='coerce').fillna(0).apply(lambda x: round(x, 2))
    except Exception as e:
        print("Error al convertir 'PF (Precio Final)':", e)
    df['Inventario Inicial'] = pd.to_numeric(df['Inventario Inical'], errors='coerce').fillna(0)
    return df[['Codigo', 'Descripcion', 'Unidades de Presentacion', 'PF (Precio Final)', 'Inventario Inicial']]

def guardar_movimiento(tipo, codigo, descripcion, unidades, precio, cantidad):
    hoja = "Entradas" if tipo == "entrada" else "Salidas"
    try:
        book = load_workbook(EXCEL_FILE)
        if hoja not in book.sheetnames:
            sheet = book.create_sheet(hoja)
            sheet.append(["Fecha", "Codigo", "Descripcion", "Unidades", "Precio", "Cantidad"])
        else:
            sheet = book[hoja]
        fecha_movimiento = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nueva_fila = [fecha_movimiento, codigo, descripcion, unidades, precio, cantidad]
        sheet.append(nueva_fila)
        book.save(EXCEL_FILE)
        messagebox.showinfo("Éxito", f"{tipo.capitalize()} registrada correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el movimiento: {str(e)}")

def cargar_movimientos(tipo):
    hoja = "Entradas" if tipo == "entrada" else "Salidas"
    try:
        return pd.read_excel(EXCEL_FILE, sheet_name=hoja)
    except Exception:
        return pd.DataFrame(columns=["Fecha", "Codigo", "Descripcion", "Unidades", "Precio", "Cantidad"])

def configurar_tabla(frame, columns):
    tabla = ttk.Treeview(frame, columns=columns, show="headings")
    for col in columns:
        tabla.heading(col, text=col)
    tabla.pack(pady=10, fill="both", expand=True)
    return tabla

def boton_regresar(frame, target_frame):
    return tk.Button(frame, text="Regresar al Menú Principal", command=lambda: [frame.pack_forget(), target_frame.pack(fill="both", expand=True)], font=("Helvetica", 12, "bold"), bg="#4285f4", fg="white")

def mostrar_reporte(tipo, frame, frame_menu):
    hoja = "Entradas" if tipo == "entrada" else "Salidas"
    df_reporte = cargar_movimientos(tipo)
    for widget in frame.winfo_children():
        widget.destroy()

    if df_reporte.empty:
        messagebox.showinfo("Sin datos", f"No hay datos registrados en {hoja}.")
        boton_regresar(frame, frame_menu).pack(pady=20)
        return

    tk.Label(frame, text=f"Reporte de {hoja}", font=("Helvetica", 14)).pack(pady=10)
    tabla = configurar_tabla(frame, ["Fecha", "Codigo", "Descripcion", "Unidades", "Precio", "Cantidad"])
    for _, row in df_reporte.iterrows():
        tabla.insert("", tk.END, values=tuple(row))
    boton_regresar(frame, frame_menu).pack(pady=20)

def mostrar_estado_inventario(frame, frame_menu):
    df_catalogo = cargar_catalogo()
    df_entradas = cargar_movimientos("entrada").groupby("Codigo")["Cantidad"].sum()
    df_salidas = cargar_movimientos("salida").groupby("Codigo")["Cantidad"].sum()

    df_catalogo["Entradas"] = df_catalogo["Codigo"].map(df_entradas).fillna(0)
    df_catalogo["Salidas"] = df_catalogo["Codigo"].map(df_salidas).fillna(0)
    df_catalogo["Inventario Actual"] = df_catalogo["Inventario Inicial"] + df_catalogo["Entradas"] - df_catalogo["Salidas"]

    # Filtrar productos con inventario actual diferente de 0
    df_catalogo = df_catalogo[df_catalogo["Inventario Actual"] != 0]

    for widget in frame.winfo_children():
        widget.destroy()

    tk.Label(frame, text="Estado del Inventario", font=("Helvetica", 14)).pack(pady=10)
    tabla = configurar_tabla(frame, ["Codigo", "Descripcion", "Unidades de Presentacion", "PF (Precio Final)", "Inventario Actual"])
    for _, row in df_catalogo.iterrows():
        tabla.insert("", tk.END, values=(
            row["Codigo"],
            row["Descripcion"],
            row["Unidades de Presentacion"],
            f"${row['PF (Precio Final)']:,.2f}",
            f"{row['Inventario Actual']:.2f}"
        ))
    boton_regresar(frame, frame_menu).pack(pady=20)

def crear_frame_registro(frame, titulo, tipo, callback_actualizar, frame_menu):
    tk.Label(frame, text=titulo, font=("Helvetica", 14)).pack(pady=10)
    entry_buscar = tk.Entry(frame, width=40)
    entry_buscar.pack(pady=5)
    tabla = configurar_tabla(frame, ["Codigo", "Descripcion", "Unidades de Presentacion", "PF (Precio Final)"])
    entry_buscar.bind("<KeyRelease>", lambda e: callback_actualizar(entry_buscar.get().lower(), tabla))

    tk.Label(frame, text="Cantidad:", font=("Helvetica", 12)).pack(pady=5)
    entry_cantidad = tk.Entry(frame, width=20)
    entry_cantidad.pack(pady=5)

    tk.Button(frame, text=f"Registrar {tipo.capitalize()}", font=("Helvetica", 12), bg="#34A853", fg="white",
              command=lambda: registrar_movimiento(tipo, entry_cantidad.get(), tabla)).pack(pady=10)
    boton_regresar(frame, frame_menu).pack(pady=20)
    return entry_buscar, entry_cantidad, tabla

def registrar_movimiento(tipo, cantidad, tabla):
    if not cantidad.isdigit():
        messagebox.showerror("Error", "Ingrese una cantidad válida.")
        return
    selected_item = tabla.focus()
    if not selected_item:
        messagebox.showerror("Error", "Seleccione un producto de la tabla.")
        return
    item_data = tabla.item(selected_item, 'values')
    guardar_movimiento(tipo, item_data[0], item_data[1], item_data[2], item_data[3], int(cantidad))
    messagebox.showinfo("Éxito", f"{tipo.capitalize()} registrada correctamente.")

def actualizar_tabla(filtro, tabla, df):
    resultados = df[(df['Descripcion'].str.lower().str.contains(filtro)) | (df['Codigo'].str.lower().str.contains(filtro))]
    tabla.delete(*tabla.get_children())
    for _, row in resultados.iterrows():
        tabla.insert("", tk.END, values=tuple(row))

def ventana_principal():
    verificar_archivo_excel()
    ventana = tk.Tk()
    ventana.title("Sistema de Inventario")
    ventana.geometry("1200x700")

    # DataFrame inicial
    df_catalogo = cargar_catalogo()

    # Frames principales
    frame_menu = tk.Frame(ventana)
    frame_entradas = tk.Frame(ventana)
    frame_salidas = tk.Frame(ventana)
    frame_reporte_entradas = tk.Frame(ventana)
    frame_reporte_salidas = tk.Frame(ventana)
    frame_inventario = tk.Frame(ventana)

    # Menú principal
    tk.Label(frame_menu, text="Menú Principal", font=("Helvetica", 16, "bold")).pack(pady=20)
    tk.Button(frame_menu, text="Entradas", command=lambda: [frame_menu.pack_forget(), frame_entradas.pack(fill="both", expand=True)]).pack(pady=10)
    tk.Button(frame_menu, text="Salidas", command=lambda: [frame_menu.pack_forget(), frame_salidas.pack(fill="both", expand=True)]).pack(pady=10)
    tk.Button(frame_menu, text="Reporte de Entradas", command=lambda: [frame_menu.pack_forget(), frame_reporte_entradas.pack(fill="both", expand=True), mostrar_reporte("entrada", frame_reporte_entradas, frame_menu)]).pack(pady=10)
    tk.Button(frame_menu, text="Reporte de Salidas", command=lambda: [frame_menu.pack_forget(), frame_reporte_salidas.pack(fill="both", expand=True), mostrar_reporte("salida", frame_reporte_salidas, frame_menu)]).pack(pady=10)
    tk.Button(frame_menu, text="Estado del Inventario", command=lambda: [frame_menu.pack_forget(), frame_inventario.pack(fill="both", expand=True), mostrar_estado_inventario(frame_inventario, frame_menu)]).pack(pady=10)
    frame_menu.pack(fill="both", expand=True)

    # Frame Entradas
    def actualizar_entradas_tabla(filtro, tabla):
        actualizar_tabla(filtro, tabla, df_catalogo)

    tk.Label(frame_entradas, text="Registro de Entradas", font=("Helvetica", 14)).pack(pady=10)
    entry_buscar_entradas, entry_cantidad_entradas, tabla_entradas = crear_frame_registro(frame_entradas, "Entradas", "entrada", actualizar_entradas_tabla, frame_menu)
    actualizar_entradas_tabla("", tabla_entradas)

    # Frame Salidas
    def actualizar_salidas_tabla(filtro, tabla):
        actualizar_tabla(filtro, tabla, df_catalogo)

    tk.Label(frame_salidas, text="Registro de Salidas", font=("Helvetica", 14)).pack(pady=10)
    entry_buscar_salidas, entry_cantidad_salidas, tabla_salidas = crear_frame_registro(frame_salidas, "Salidas", "salida", actualizar_salidas_tabla, frame_menu)
    actualizar_salidas_tabla("", tabla_salidas)

    ventana.mainloop()

if __name__ == "__main__":
    ventana_principal()
